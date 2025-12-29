from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List
from src.core.models import ClaseConDia

class ExcelExporter:
    """Servicio encargado de renderizar y exportar horarios a formato Excel Premium"""
    
    COLORES_EXCEL = ["BFDBFE", "BBF7D0", "FEF3C7", "FECACA", "DDD6FE", "F5D0FE", "FED7AA"]
    DIAS_SEMANA = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']

    @staticmethod
    def exportar(clases: List[ClaseConDia], nombre_archivo: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Horario Optimizado"
        
        # --- TABLA RESUMEN ---
        headers = ["NRC", "Asignatura", "Tipo", "Sección", "Día", "Horario", "Docente"]
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin = Side(style='thin')
        header_border = Border(bottom=thin)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font, cell.fill = header_font, header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = header_border
            ws.column_dimensions[get_column_letter(col)].width = 16
        
        ws.column_dimensions['B'].width = 34
        
        # Llenar datos de ramos (tabla resumen)
        for i, c in enumerate(clases, 2):
            ws.cell(row=i, column=1, value=c.nrc)
            ws.cell(row=i, column=2, value=c.titulo)
            ws.cell(row=i, column=3, value=c.tipo)
            ws.cell(row=i, column=4, value=c.seccion)
            ws.cell(row=i, column=5, value=c.dia)
            ws.cell(row=i, column=6, value=f"{c.hora_inicio} - {c.hora_fin}")
            ws.cell(row=i, column=7, value=c.instructor)
            
            for col in range(1, 8):
                ws.cell(row=i, column=col).border = Border(bottom=thin)

        # --- GRILLA SEMANAL ---
        start_row = len(clases) + 5
        ws.cell(row=start_row-1, column=1, value="MAPA SEMANAL").font = Font(size=14, bold=True)
        # Freeze header rows for easier navigation
        ws.freeze_panes = ws['A2']
        
        # Headers de días
        for i, dia in enumerate(ExcelExporter.DIAS_SEMANA):
            cell = ws.cell(row=start_row, column=2 + i, value=dia)
            cell.font, cell.alignment = Font(bold=True), Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            cell.border = Border(bottom=thin)
            ws.column_dimensions[get_column_letter(2+i)].width = 20

        # Filas de horas
        row_map = {}
        for i, h in enumerate(range(8, 23)):
            row = start_row + 1 + i
            ws.cell(row=row, column=1, value=f"{h:02d}:00").font = Font(bold=True)
            row_map[h] = row
            for d in range(6):
                ws.cell(row=row, column=2+d).border = Border(left=Side(style='dotted'), bottom=Side(style='dotted'))

        # Bloquear celdas con ramos
        # Asignar color base por título y tono según tipo (TEO claro, LAB oscuro)
        def hex_to_rgb(h: str):
            h = h.lstrip('#')
            return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

        def rgb_to_hex(rgb):
            return '%02X%02X%02X' % (max(0, min(255, int(rgb[0]))), max(0, min(255, int(rgb[1]))), max(0, min(255, int(rgb[2]))))

        def adjust_brightness(hexcol: str, factor: float):
            r, g, b = hex_to_rgb(hexcol)
            return rgb_to_hex((r * factor, g * factor, b * factor))

        title_base = {}
        nrc_colors = {}
        idx = 0
        for c in clases:
            if c.titulo not in title_base:
                base = ExcelExporter.COLORES_EXCEL[idx % len(ExcelExporter.COLORES_EXCEL)]
                title_base[c.titulo] = base
                idx += 1

        for c in clases:
            base = title_base.get(c.titulo, ExcelExporter.COLORES_EXCEL[0])
            base_hex = base if base.startswith('#') else f"#{base}"
            tipo_up = (c.tipo or '').upper()
            tipo_norm = 'TEO' if ('TEOR' in tipo_up or 'TEO' in tipo_up) else 'LAB' if ('LAB' in tipo_up or 'TALLER' in tipo_up or 'PRACT' in tipo_up) else 'OTRO'
            
            if tipo_norm == 'TEO':
                tone = adjust_brightness(base_hex, 1.15)
            elif tipo_norm == 'LAB':
                tone = adjust_brightness(base_hex, 0.82)
            else:
                tone = base_hex.lstrip('#')
            
            if c.nrc not in nrc_colors:
                nrc_colors[c.nrc] = tone

        for c in clases:
            color_hex = nrc_colors[c.nrc].lstrip('#')
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            try:
                h_start = int(c.hora_inicio.split(':')[0])
                h_end = int(c.hora_fin.split(':')[0])
                if c.dia in ExcelExporter.DIAS_SEMANA:
                    col_idx = 2 + ExcelExporter.DIAS_SEMANA.index(c.dia)
                    # Rango inclusivo (si termina a hh:30 contamos la hora final)
                    rango = range(h_start, h_end + (1 if int(c.hora_fin.split(':')[1]) > 0 else 0))
                    rows_present = [row_map[h] for h in rango if h in row_map]
                    if rows_present:
                        r0, r1 = rows_present[0], rows_present[-1]
                        # Pintar y, si ocupa más de 1 hora, mergear para una apariencia limpia
                        for r in rows_present:
                            cell = ws.cell(row=r, column=col_idx)
                            cell.fill = fill
                            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        # Texto en la celda superior (merge)
                        try:
                            if r1 > r0:
                                ws.merge_cells(start_row=r0, start_column=col_idx, end_row=r1, end_column=col_idx)
                            top_cell = ws.cell(row=r0, column=col_idx)
                            tipo_up = (c.tipo or '').upper()
                            tipo_norm = 'TEO' if ('TEOR' in tipo_up or 'TEO' in tipo_up) else 'LAB' if ('LAB' in tipo_up or 'TALLER' in tipo_up or 'PRACT' in tipo_up) else 'OTRO'
                            top_cell.value = f"{c.titulo} ({tipo_norm})\n({c.nrc})"
                            top_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                            # Decide font color depending on luminance
                            r_col = int(color_hex[0:2], 16); g_col = int(color_hex[2:4], 16); b_col = int(color_hex[4:6], 16)
                            luminance = 0.2126*r_col + 0.7152*g_col + 0.0722*b_col
                            top_cell.font = Font(bold=True, color="FFFFFF" if luminance < 150 else "000000")
                        except Exception:
                            pass
            except: pass
        # Añadir leyenda de colores a la derecha de la tabla resumen
        try:
            legend_col = 9
            legend_row = 2
            ws.cell(row=legend_row-1, column=legend_col, value="Leyenda").font = Font(bold=True)
            for idx, (titulo, hexcol) in enumerate(title_base.items(), start=0):
                r = legend_row + idx
                col = legend_col
                ccell = ws.cell(row=r, column=col, value=titulo)
                try:
                    fill_color = title_base[titulo].lstrip('#')
                    ws.cell(row=r, column=col+1).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    ws.cell(row=r, column=col+1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                except Exception:
                    pass
                ccell.alignment = Alignment(vertical='center')
        except Exception:
            pass

        # Auto-filter y propiedades
        try:
            last_row = len(clases) + 1
            ws.auto_filter.ref = f"A1:G{last_row}"
        except Exception:
            pass

        wb.properties.title = "Horario Optimizado"
        wb.save(nombre_archivo)